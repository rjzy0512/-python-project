# TODO 导入openpyxl模块

import openpyxl
# TODO 使用openpyxl.Workbook()函数，创建一个新工作薄，赋值给newW=
newW=openpyxl.Workbook()

# TODO 将默认的工作表Sheet赋值给aSheet=变量

aSheet=newW["Sheet"]
# TODO 将aSheet工作表名称修改为“七年级1班课表”
aSheet.title="七年级1班课表"

# TODO 使用for循环和range，逐个遍历2~10的数字
for xunhuan in range(2,11):
    # TODO 利用格式化字符串"七年级x班课表",创建不同班级的工作表
    newW.create_sheet(f"七年级{xunhuan}班课表")

# TODO 使用for循环遍历工作簿对象的worksheets属性
for biaotou in newW.worksheets:

    # TODO 给每一个工作表设置表头，B1单元格设置为星期一，C1单元格设置为星期二
    biaotou["c1"].value = "星期二"
    biaotou["B1"].value = "星期一"
    biaotou["D1"].value = "星期三"
    biaotou["E1"].value = "星期四"
    biaotou["F1"].value = "星期五"

    # TODO 使用merge_cells()函数，合并单元格A2:A4
    aSheet.merge_cells('A2:A4')
    # TODO 将单元格A2的值设置为上午
    biaotou["A2"].value = "上午"

    # TODO 使用merge_cells()函数，合并单元格A5:A6
    aSheet.merge_cells('A5:A6')
    # TODO 将单元格A5的值设置为下午
    biaotou["A5"].value = "下午"

# TODO 将工作簿保存到路径 /Users/yequ/Desktop/七年级课表.xlsx
newW.save("/Users/yequ/Desktop/七年级课表.xlsx")
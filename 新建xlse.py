# 导入openpyxl 模块
import openpyxl

# TODO 使用openpyxl.Workbook()函数创建一个新工作簿并赋值给变量newWb=
newWb=openpyxl.Workbook()

# TODO 将默认工作表Sheet赋值给aSheet=变量

aSheet=newWb["Sheet"]
# TODO 将aSheet工作表名称修改为“东胜神洲”
aSheet.title="东胜神洲"

# TODO 创建 西牛贺洲 的工作表
newWb.create_sheet("西牛贺洲")
# TODO 创建 南赡部洲 的工作表
newWb.create_sheet("南赡部洲")
# TODO 创建 北俱芦洲 的工作表
newWb.create_sheet("北俱芦洲")

# TODO 使用for循环遍历工作簿对象的worksheets属性
for sheet in newWb.worksheets:
    # 给每一个工作表设置表头
    # TODO 将A1单元格的值设置为 怪物编号
    sheet["A1"].value="怪物编号"
    # TODO 将B1单元格的值设置为 怪物名称
    sheet["B1"].value="怪物名称"
    # TODO 将C1单元格的值设置为 怪物血量
    sheet["C1"].value="怪物血量"

# TODO 将工作簿保存到路径 /Users/yequ/data/怪物数值.xlsx
newWb.save("/Users/yequ/data/怪物数值.xlsx")
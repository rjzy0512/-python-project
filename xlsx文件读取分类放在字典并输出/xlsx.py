# 导入openpyxl 模块
import openpyxl

# 创建一个新工作簿
newWb = openpyxl.Workbook()

# 将默认工作表赋值给aSheet变量
aSheet = newWb["Sheet"]

# 将aSheet工作表名称修改为“A平台”
aSheet.title = "A平台"

# 创建 B平台 的工作表并赋值给变量bSheet
bSheet = newWb.create_sheet("B平台")

# 创建 C平台 的工作表并赋值给变量cSheet
cSheet = newWb.create_sheet("C平台")

# 使用for循环遍历工作簿对象的worksheets属性
for sheet in newWb.worksheets:
    # 给每一个工作表设置表头
    sheet["A1"].value = "商品名"
    sheet["B1"].value = "月份"
    sheet["C1"].value = "销售额"


# 计算月份工作簿内各个商品的总价,并添加数据到汇总工作表
# 参数 filePath: 工作簿文件路径
# 参数 orderSheetName: 订单工作表的名称
# 参数 nameIndex: 商品名称的列索引
# 参数 nameHead: 商品名称的表头
# 参数 priceColumn: 总价的列号
# 参数 month: 当前处理的月份
# 参数 targetSheet: 要添加到的目标工作表
def processMonthFile(filePath, orderSheetName, nameIndex, nameHead, priceColumn, month, targetSheet):
    # 使用load_wordbook函数读取工作簿
    wb = openpyxl.load_workbook(filePath, data_only=True)
    # 获取订单工作表
    orderSheet = wb[orderSheetName]

    # 定义一个空字典用来放本月数据
    monthData = {}

    # 遍历订单工作表，计算每个商品的总销售额
    for row in orderSheet.rows:
        # 获取订单商品名称
        productName = row[nameIndex].value
        # 跳过表头
        if productName == nameHead:
            continue
        # 获取总价列的索引
        priceIndex = openpyxl.utils.cell.column_index_from_string(priceColumn) - 1
        # 获取订单总价
        orderPrice = row[priceIndex].value

        # TODO 先安全地获取已经统计到的商品月销售额赋值给monthPrice
        # 如果productName还不在字典monthData中，返回0
        monthPrice = monthData.get(productName, 0)
        # 将这个订单的价格累加到商品总价
        monthData[productName] = monthPrice + orderPrice

    # 遍历本月数据字典的keys，也就是商品名称
    for productName in monthData.keys():
        # 按顺序构造行数据元组: 商品名, 月份, 销售额数据
        rowData = (productName, f"{month}月份", monthData[productName])
        # TODO 使用append()函数将rowData添加到目标工作表targetSheet中
        targetSheet.append(rowData)


# 使用for循环和range，逐个遍历1~12的数字
for month in range(1, 13):
    # 使用processMonthFile函数，汇总A平台数据
    # TODO 格式化输出A平台文件路径：A平台/2019XX.xlsx，其中XX是月份
    processMonthFile(f"A平台/2019{month:02}.xlsx", "明细", 4, "名称", "F", month, aSheet)
    # 汇总B平台数据
    processMonthFile(f"B平台/order_2019_{month}.xlsx", "订单详情", 1, "商品名称", "G", month, bSheet)
    # 汇总C平台数据
    processMonthFile(f"C平台/2019年{month}月销售订单.xlsx", "销售订单数据", 2, "商品名", "I", month, cSheet)

# 将工作簿保存到路径 /Users/yequ/data/汇总.xlsx
newWb.save("/Users/yequ/data/汇总.xlsx")